import logging
import tempfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from .heatmap import add_heatmap
from .pairplot import add_pairplot
from .robustness import add_robustness

logger = logging.getLogger(__name__)


def add_plots_to_excel(excel_path, results_df, config):
    """
    Tüm grafikleri oluşturup Excel dosyasına ekler.
    
    Args:
        excel_path: Hedef Excel dosyası yolu
        results_df: Araştırma sonuçlarını içeren DataFrame
        config: ResearchConfig nesnesi
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        img_files = []
        img_files.extend(add_heatmap(results_df, tmpdir))
        img_files.extend(add_pairplot(results_df, config, tmpdir))
        img_files.extend(add_robustness(results_df, tmpdir))
        
        if not img_files:
            logger.warning("Hiç görsel oluşturulamadı.")
            return
        
        try:
            wb = load_workbook(excel_path)
            for sheet_name, img_file in img_files:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                img = Image(img_file)
                img.width = 800
                img.height = 600
                ws.add_image(img, 'A1')
            wb.save(excel_path)
            logger.info(f"🖼️ {len(img_files)} görsel Excel'e eklendi.")
        except Exception as e:
            logger.error(f"Görseller Excel'e eklenemedi: {e}")