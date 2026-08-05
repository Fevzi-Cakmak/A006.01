# research/reporting/reporting.py
"""
Raporlama Modülü

- ExcelReporter: Excel dosyası oluşturur (stillendirmeli)
- HTMLReporter: Plotly ile interaktif HTML raporu
- SummaryReporter: Özet istatistikler
"""

import os
import json
import warnings
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import pandas as pd
import numpy as np


class ExcelReporter:
    """
    Excel raporlayıcı.
    Birden fazla sayfayı stillendirilmiş olarak yazar.
    """

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.sheets: List[Tuple[pd.DataFrame, str, bool]] = []

    def add_sheet(self, df: pd.DataFrame, name: str, index: bool = False) -> "ExcelReporter":
        self.sheets.append((df, name, index))
        return self

    def save(self, engine: str = "openpyxl") -> str:
        """Excel dosyasını kaydeder."""
        if not self.sheets:
            raise ValueError("Hiç sayfa eklenmemiş.")

        with pd.ExcelWriter(self.filepath, engine=engine) as writer:
            for df, name, index in self.sheets:
                df.to_excel(writer, sheet_name=name, index=index)

        # Stil ekle (openpyxl ile)
        try:
            self._apply_styles()
        except Exception as e:
            warnings.warn(f"Stil uygulanamadı: {e}")

        return self.filepath

    def _apply_styles(self):
        """Openpyxl ile stil ekler."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            return

        wb = load_workbook(self.filepath)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        alt_fill = PatternFill("solid", fgColor="D9EAF7")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for ws in wb.worksheets:
            # Başlıklar
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Alternatif satır renkleri
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for cell in row:
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            # Sütun genişlikleri
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

        wb.save(self.filepath)


class HTMLReporter:
    """
    HTML raporlayıcı (Plotly ile interaktif).
    """

    def __init__(self, filepath: str, title: str = "Araştırma Raporu"):
        self.filepath = filepath
        self.title = title
        self.sections: List[Dict[str, Any]] = []

    def add_text(self, content: str, level: str = "h2") -> "HTMLReporter":
        self.sections.append({"type": "text", "level": level, "content": content})
        return self

    def add_table(self, df: pd.DataFrame, caption: str = "") -> "HTMLReporter":
        self.sections.append({"type": "table", "caption": caption, "data": df})
        return self

    def add_plotly(self, fig, caption: str = "") -> "HTMLReporter":
        """Plotly figure ekler."""
        self.sections.append({"type": "plotly", "caption": caption, "fig": fig})
        return self

    def add_image(self, filepath: str, caption: str = "") -> "HTMLReporter":
        self.sections.append({"type": "image", "caption": caption, "path": filepath})
        return self

    def save(self) -> str:
        """HTML dosyasını kaydeder."""
        html = self._render_html()
        with open(self.filepath, "w", encoding="utf-8") as f:
            f.write(html)
        return self.filepath

    def _render_html(self) -> str:
        """HTML içeriğini oluşturur."""
        # Plotly varsa import et
        try:
            import plotly.io as pio
            HAS_PLOTLY = True
        except ImportError:
            HAS_PLOTLY = False

        sections_html = []
        for section in self.sections:
            if section["type"] == "text":
                level = section.get("level", "h2")
                sections_html.append(f"<{level}>{section['content']}</{level}>")
            elif section["type"] == "table":
                df = section["data"]
                caption = section.get("caption", "")
                table_html = df.to_html(classes="dataframe", border=0)
                if caption:
                    table_html = f"<caption>{caption}</caption>" + table_html
                sections_html.append(table_html)
            elif section["type"] == "plotly":
                if HAS_PLOTLY:
                    fig = section["fig"]
                    plot_html = pio.to_html(fig, full_html=False, include_plotlyjs="cdn")
                    caption = section.get("caption", "")
                    if caption:
                        plot_html = f"<p><em>{caption}</em></p>" + plot_html
                    sections_html.append(plot_html)
                else:
                    sections_html.append("<p>Plotly kurulu değil.</p>")
            elif section["type"] == "image":
                path = section["path"]
                caption = section.get("caption", "")
                if os.path.exists(path):
                    img_html = f'<img src="{path}" style="max-width:100%;"/>'
                    if caption:
                        img_html = f"<p><em>{caption}</em></p>" + img_html
                    sections_html.append(img_html)
                else:
                    sections_html.append(f"<p>Görsel bulunamadı: {path}</p>")

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>{self.title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background: #f8f9fa; }}
                .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                h1, h2, h3 {{ color: #1F4E78; }}
                .dataframe {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                .dataframe th {{ background: #1F4E78; color: white; padding: 10px; text-align: center; }}
                .dataframe td {{ padding: 8px; text-align: center; border-bottom: 1px solid #ddd; }}
                .dataframe tr:nth-child(even) {{ background: #f2f2f2; }}
                img {{ max-width: 100%; height: auto; }}
                caption {{ font-weight: bold; margin: 10px 0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>{self.title}</h1>
                <p><small>Oluşturulma: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</small></p>
                <hr>
                {''.join(sections_html)}
            </div>
        </body>
        </html>
        """
        return html


class SummaryReporter:
    """
    Özet istatistik raporu.
    """

    def __init__(self):
        self.summary: Dict[str, Any] = {}

    def add_metric(self, name: str, value: Any, description: str = "") -> "SummaryReporter":
        self.summary[name] = {"value": value, "description": description}
        return self

    def add_dataframe_stats(self, df: pd.DataFrame, prefix: str = "") -> "SummaryReporter":
        """DataFrame istatistiklerini ekler."""
        if prefix:
            prefix = prefix + "_"
        self.summary[f"{prefix}shape"] = {"value": df.shape, "description": "Boyut"}
        self.summary[f"{prefix}columns"] = {"value": list(df.columns), "description": "Sütunlar"}
        if not df.empty:
            self.summary[f"{prefix}null"] = {"value": df.isnull().sum().sum(), "description": "Toplam boş değer"}
        return self

    def to_dataframe(self) -> pd.DataFrame:
        """Özeti DataFrame'e dönüştürür."""
        rows = []
        for name, data in self.summary.items():
            rows.append({
                "Metrik": name,
                "Değer": data["value"],
                "Açıklama": data.get("description", "")
            })
        return pd.DataFrame(rows)

    def to_dict(self) -> Dict[str, Any]:
        return {k: v["value"] for k, v in self.summary.items()}

    def print_summary(self):
        for name, data in self.summary.items():
            print(f"{name}: {data['value']}")


class ReportFactory:
    """Rapor oluşturma fabrikası."""

    @staticmethod
    def create_excel(filepath: str) -> ExcelReporter:
        return ExcelReporter(filepath)

    @staticmethod
    def create_html(filepath: str, title: str = "Araştırma Raporu") -> HTMLReporter:
        return HTMLReporter(filepath, title)

    @staticmethod
    def create_summary() -> SummaryReporter:
        return SummaryReporter()