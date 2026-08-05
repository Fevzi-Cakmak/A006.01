# core.py
# A003.3 - Portfolio modüler, tüm fonksiyonlar ve sınıflar eksiksiz.

from __future__ import annotations

import logging
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Optional, Protocol, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

logger = logging.getLogger(__name__)

try:
    from google.colab import files

    COLAB = True
except ImportError:
    COLAB = False


# =========================================================
# 1. KONFIGÜRASYON
# =========================================================
@dataclass
class BacktestConfig:
    backtest_start: str = "2021-01-01"
    backtest_end: Optional[str] = None
    data_start: str = "2020-01-01"
    min_data_count: int = 220

    strong_buy_min_score: int = 90
    volume_high_ratio: float = 2.0
    volume_upper_ratio: float = 4.0
    adx_strong: int = 30
    adx_upper: int = 45
    rsi_low: int = 58
    rsi_high: int = 65

    stop_loss_ratio: float = 0.1025
    trailing_start: float = 0.25
    trailing_stop_ratio: float = 0.07
    max_hold_days: int = 120
    partial_profit_threshold: float = 0.18
    partial_profit_ratio: float = 0.50

    kademeli_satis_plani: List[Tuple[float, float]] = field(
        default_factory=lambda: [(0.15, 0.35), (0.22, 0.25), (0.29, 0.15)]
    )

    initial_capital: float = 100_000.0
    max_positions: int = 5
    first_entry_ratio: float = 0.10
    add_entry_ratio: float = 0.89
    max_single_stock_ratio: float = 0.98
    profit_trigger_pct: float = 8.0

    commission: float = 0.0010
    slippage: float = 0.0010

    evren_endeks: str = "XUTUM"
    output_dir: str = "VIOS_A004_PYR5_2D_8_Portfoy"
    giris_sinyalleri: Set[str] = field(
        default_factory=lambda: {"GÜÇLÜ AL (ONAYLI) 🚀"}
    )


config = BacktestConfig()


# =========================================================
# 2. DATA FETCHER PROTOCOL
# =========================================================
class DataFetcherProtocol(Protocol):
    def get_ohlcv(
        self, symbol: str, start: Optional[str] = None, end: Optional[str] = None
    ) -> pd.DataFrame:
        ...

    def clear_cache(self) -> None:
        ...


# =========================================================
# 3. İNDİKATÖR FONKSİYONLARI (A003.2 optimizasyonu ile)
# =========================================================
def rsi_wilder(series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))


def bollinger(series, window=20, num_std=2):
    mid = series.rolling(window).mean()
    std = series.rolling(window).std(ddof=0)
    return mid + num_std * std, mid, mid - num_std * std


def true_range(df):
    """pd.concat yerine np.maximum ile çok daha hızlı True Range hesaplama."""
    prev_close = df["Close"].shift(1)
    tr1 = df["High"] - df["Low"]
    tr2 = (df["High"] - prev_close).abs()
    tr3 = (df["Low"] - prev_close).abs()
    return np.maximum(tr1, np.maximum(tr2, tr3))


def atr_wilder(df, period=14):
    return true_range(df).ewm(alpha=1 / period, adjust=False, min_periods=period).mean()


def adx_wilder(df, period=14):
    high_diff = df["High"].diff()
    low_diff = -df["Low"].diff()
    plus_dm = pd.Series(
        np.where((high_diff > low_diff) & (high_diff > 0), high_diff, 0.0),
        index=df.index,
    )
    minus_dm = pd.Series(
        np.where((low_diff > high_diff) & (low_diff > 0), low_diff, 0.0),
        index=df.index,
    )
    atr = atr_wilder(df, period)
    plus_di = (
        100 * plus_dm.ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atr
    )
    minus_di = (
        100 * minus_dm.ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atr
    )
    denom = (plus_di + minus_di).replace(0, np.nan)
    dx = 100 * (plus_di - minus_di).abs() / denom
    adx = dx.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    return adx, plus_di, minus_di


def compute_ema(series, span):
    return series.ewm(span=span, adjust=False).mean()


def compute_sma(series, window):
    return series.rolling(window).mean()


def compute_volume_ratio(df):
    """replace yerine where kullanarak hızlandırıldı."""
    vol_avg = df["Volume"].rolling(20).mean()
    vol_avg = vol_avg.where(vol_avg != 0, np.nan)
    return df["Volume"] / vol_avg


def compute_previous_20_day_high(df):
    return df["High"].rolling(20).max().shift(1)


# =========================================================
# 4. VEKTÖREL SKOR VE SİNYAL
# =========================================================
def hesapla_skor_vektorel(df: pd.DataFrame, cfg: BacktestConfig) -> np.ndarray:
    close = df["Close"].to_numpy()
    sma20 = df["SMA20"].to_numpy()
    sma50 = df["SMA50"].to_numpy()
    sma200 = df["SMA200"].to_numpy()
    sma200_egim = df["SMA200_Egim"].to_numpy()
    rsi = df["RSI"].to_numpy()
    macd = df["MACD"].to_numpy()
    macd_signal = df["MACD_Signal"].to_numpy()
    adx = df["ADX"].to_numpy()
    hacim = df["Hacim_Orani"].to_numpy()
    onceki_zirve = df["Onceki_20G_Zirve"].to_numpy()
    bb_upper = df["BB_Upper"].to_numpy()
    ema12 = df["EMA12"].to_numpy()
    ema26 = df["EMA26"].to_numpy()

    skor = np.zeros(len(df), dtype=np.int16)

    skor += np.select(
        [(close > sma200) & (sma200_egim > 0), close > sma200],
        [12, 6],
        default=0,
    ).astype(np.int16)

    skor += np.where(sma50 > sma200, 10, 0).astype(np.int16)
    skor += np.where(sma20 > sma50, 8, 0).astype(np.int16)

    skor += np.select(
        [
            (rsi >= cfg.rsi_low) & (rsi <= cfg.rsi_high),
            (rsi >= 50) & (rsi < cfg.rsi_low),
        ],
        [12, 6],
        default=0,
    ).astype(np.int16)

    skor += np.where(macd > macd_signal, 8, 0).astype(np.int16)

    skor += np.select(
        [adx >= 40, adx >= 35, adx >= 30, adx >= 25],
        [20, 17, 14, 8],
        default=0,
    ).astype(np.int16)

    skor += np.select(
        [hacim >= 3.0, hacim >= 2.5, hacim >= 2.0, hacim >= 1.5],
        [15, 12, 9, 5],
        default=0,
    ).astype(np.int16)

    fiyat_skor = np.select(
        [close > onceki_zirve, close >= bb_upper * 0.98],
        [8, 4],
        default=0,
    ).astype(np.int16)

    fiyat_skor += np.select(
        [(close > ema12) & (ema12 > ema26), ema12 > ema26],
        [7, 3],
        default=0,
    ).astype(np.int16)

    skor += np.minimum(fiyat_skor, 15).astype(np.int16)
    return np.clip(skor, 0, 100).astype(np.int16)


def hesapla_sinyal_vektorel(df: pd.DataFrame, cfg: BacktestConfig) -> np.ndarray:
    close = df["Close"].to_numpy()
    open_ = df["Open"].to_numpy()
    sma20 = df["SMA20"].to_numpy()
    sma50 = df["SMA50"].to_numpy()
    sma200 = df["SMA200"].to_numpy()
    sma200_egim = df["SMA200_Egim"].to_numpy()
    ema12 = df["EMA12"].to_numpy()
    ema26 = df["EMA26"].to_numpy()
    rsi = df["RSI"].to_numpy()
    macd = df["MACD"].to_numpy()
    macd_signal = df["MACD_Signal"].to_numpy()
    adx = df["ADX"].to_numpy()
    hacim = df["Hacim_Orani"].to_numpy()
    plus_di = df["Plus_DI"].to_numpy()
    minus_di = df["Minus_DI"].to_numpy()
    onceki_zirve = df["Onceki_20G_Zirve"].to_numpy()
    bb_upper = df["BB_Upper"].to_numpy()
    bb_lower = df["BB_Lower"].to_numpy()
    bb_genislik = df["BB_Genislik_%"].to_numpy()
    gunluk_degisim = df["Gunluk_Degisim_%"].to_numpy()
    ma_fark = df["MA20_50_Fark_%"].to_numpy()
    skor = df["VIOS_Skoru"].to_numpy()

    strong_buy = (
        (skor >= cfg.strong_buy_min_score)
        & (close > sma20)
        & (sma20 > sma50)
        & (ema12 > ema26)
        & (adx >= cfg.adx_strong)
        & (adx <= cfg.adx_upper)
        & (hacim >= cfg.volume_high_ratio)
        & ((hacim <= 3.0) | (adx >= 40))
        & (rsi >= cfg.rsi_low)
        & (rsi <= cfg.rsi_high)
        & (macd > macd_signal)
        & (plus_di > minus_di)
        & (sma200_egim > 0)
        & ((close > onceki_zirve) | (close >= bb_upper * 0.98))
    )

    sell = (
        (rsi > 72).astype(np.int8)
        + (close > bb_upper).astype(np.int8)
        + (hacim >= 2.5).astype(np.int8)
        + (gunluk_degisim >= 5.0).astype(np.int8)
    ) >= 3

    reaction_buy = (
        (rsi < 30)
        & (close <= bb_lower * 1.01)
        & (hacim >= 1.20)
        & (close > open_)
        & (close > sma200)
    )

    prep_buy = (
        (ema12 > ema26) & (rsi >= 50) & (close > sma20) & (skor >= 55)
    )

    sideways = (adx < 20) & (bb_genislik < 22) & (ma_fark < 4)

    return np.select(
        [strong_buy, sell, reaction_buy, prep_buy, sideways, rsi < 30],
        [
            "GÜÇLÜ AL (ONAYLI) 🚀",
            "SAT / KÂR AL (DOYUM)",
            "TEPKİ ALIMI (GÜVENLİ)",
            "AL (HAZIRLIK)",
            "YATAY / SIKIŞMA",
            "İZLE (AŞIRI SATIM)",
        ],
        default="NÖTR",
    )


# =========================================================
# 5. İNDİKATÖR HAZIRLAMA
# =========================================================
def indikatorleri_hazirla(df: pd.DataFrame, cfg: BacktestConfig) -> pd.DataFrame:
    df = df.sort_index()
    df = df[~df.index.duplicated(keep="last")]

    df["RSI"] = rsi_wilder(df["Close"])
    df["EMA12"] = compute_ema(df["Close"], 12)
    df["EMA26"] = compute_ema(df["Close"], 26)
    df["MACD"] = df["EMA12"] - df["EMA26"]
    df["MACD_Signal"] = compute_ema(df["MACD"], 9)

    df["SMA20"] = compute_sma(df["Close"], 20)
    df["SMA50"] = compute_sma(df["Close"], 50)
    df["SMA200"] = compute_sma(df["Close"], 200)
    df["SMA200_Egim"] = df["SMA200"].diff(5)

    df["BB_Upper"], df["BB_Mid"], df["BB_Lower"] = bollinger(df["Close"])
    df["BB_Genislik_%"] = (
        (df["BB_Upper"] - df["BB_Lower"]) / df["BB_Mid"].replace(0, np.nan) * 100
    )

    df["ADX"], df["Plus_DI"], df["Minus_DI"] = adx_wilder(df)
    df["ATR14"] = atr_wilder(df)
    df["ATR_%"] = df["ATR14"] / df["Close"].replace(0, np.nan) * 100

    df["Hacim_Orani"] = compute_volume_ratio(df)
    df["Onceki_20G_Zirve"] = compute_previous_20_day_high(df)
    df["Gunluk_Degisim_%"] = df["Close"].pct_change() * 100
    df["MA20_50_Fark_%"] = (
        (df["SMA20"] - df["SMA50"]).abs() / df["SMA50"].replace(0, np.nan) * 100
    )

    df = df.dropna()
    if len(df) < 2:
        raise ValueError("Indikator hesaplamasi sonrasi yeterli veri kalmadi.")

    df["VIOS_Skoru"] = hesapla_skor_vektorel(df, cfg)
    df["Sinyal"] = hesapla_sinyal_vektorel(df, cfg)
    return df


# =========================================================
# 6. POZİSYON SINIFI
# =========================================================
class Position:
    def __init__(
        self, symbol, entry_date, entry_price, entry_row, cfg: BacktestConfig
    ):
        self.symbol = symbol
        self.entry_date = entry_date
        self.entry_price = entry_price
        self.config = cfg

        self.giris_skoru = int(entry_row["VIOS_Skoru"])
        self.giris_sinyali = entry_row["Sinyal"]
        self.giris_rsi = round(float(entry_row["RSI"]), 2)
        self.giris_adx = round(float(entry_row["ADX"]), 2)
        self.giris_hacim = round(float(entry_row["Hacim_Orani"]), 2)
        self.giris_macd = round(float(entry_row["MACD"]), 4)
        self.giris_macd_sinyal = round(float(entry_row["MACD_Signal"]), 4)
        self.giris_atr = round(float(entry_row["ATR_%"]), 2)
        self.giris_breakout = (
            "EVET" if entry_row["Close"] > entry_row["Onceki_20G_Zirve"] else "HAYIR"
        )
        self.giris_bb_yakin = (
            "EVET" if entry_row["Close"] >= entry_row["BB_Upper"] * 0.98 else "HAYIR"
        )
        self.giris_plus_di = round(float(entry_row["Plus_DI"]), 2)
        self.giris_minus_di = round(float(entry_row["Minus_DI"]), 2)
        self.giris_sma200_egim = round(float(entry_row["SMA200_Egim"]), 4)

        self.zirve = entry_price
        self.bar_sayisi = 0
        self.kalan_oran = 1.0
        self.kismi_kar_alindi = False
        self.kademeli_satis_yapilan = set()
        self.realize_net = 0.0
        self.realize_brut = 0.0
        self.acik = True

    def update(self, today_row):
        self.zirve = max(self.zirve, float(today_row["High"]))
        self.bar_sayisi += 1

    def apply_partial_profit(self, today_row):
        for kademe_no, (kar_esigi, satis_orani) in enumerate(
            self.config.kademeli_satis_plani, start=1
        ):
            if kademe_no in self.kademeli_satis_yapilan:
                continue
            if float(today_row["High"]) < self.entry_price * (1 + kar_esigi):
                continue
            if self.kalan_oran <= 0:
                break

            uygulanacak_oran = min(satis_orani, self.kalan_oran)
            satis_fiyati = self.entry_price * (1 + kar_esigi) * (
                1 - self.config.slippage
            )
            brut_part = satis_fiyati / self.entry_price - 1
            net_part = (satis_fiyati * (1 - self.config.commission)) / (
                self.entry_price * (1 + self.config.commission)
            ) - 1
            self.realize_brut += uygulanacak_oran * brut_part
            self.realize_net += uygulanacak_oran * net_part
            self.kalan_oran -= uygulanacak_oran
            self.kademeli_satis_yapilan.add(kademe_no)
            self.kismi_kar_alindi = True

    def check_exit(self, today_row, tomorrow_row):
        stop_fiyati = self.entry_price * (1 - self.config.stop_loss_ratio)
        trailing_aktif = self.zirve >= self.entry_price * (
            1 + self.config.trailing_start
        )
        trailing_fiyati = (
            self.zirve * (1 - self.config.trailing_stop_ratio) if trailing_aktif else -np.inf
        )
        aktif_stop = max(stop_fiyati, trailing_fiyati)

        exit_date = None
        exit_price = None
        exit_reason = None

        if float(today_row["Low"]) <= aktif_stop:
            exit_date = today_row.name
            exit_price = (
                float(today_row["Open"])
                if float(today_row["Open"]) <= aktif_stop
                else aktif_stop
            ) * (1 - self.config.slippage)
            exit_reason = (
                "TRAILING STOP"
                if trailing_aktif and trailing_fiyati >= stop_fiyati
                else "STOP LOSS"
            )
        elif today_row["Sinyal"] == "SAT / KÂR AL (DOYUM)":
            exit_date = tomorrow_row.name
            exit_price = float(tomorrow_row["Open"]) * (1 - self.config.slippage)
            exit_reason = "SAT / KAR AL SINYALI"
        elif float(today_row["Close"]) < float(today_row["SMA50"]) - 2 * float(
            today_row["ATR14"]
        ):
            exit_date = tomorrow_row.name
            exit_price = float(tomorrow_row["Open"]) * (1 - self.config.slippage)
            exit_reason = "SMA50 ALTI ATR (2x)"
        elif self.config.max_hold_days > 0 and self.bar_sayisi >= self.config.max_hold_days:
            exit_date = tomorrow_row.name
            exit_price = float(tomorrow_row["Open"]) * (1 - self.config.slippage)
            exit_reason = "AZAMI SURE"

        if exit_reason:
            return {"date": exit_date, "price": exit_price, "reason": exit_reason}
        return None

    def close(self, exit_date, exit_price, exit_reason):
        brut_kalan = exit_price / self.entry_price - 1
        net_kalan = (exit_price * (1 - self.config.commission)) / (
            self.entry_price * (1 + self.config.commission)
        ) - 1
        brut_getiri = self.realize_brut + self.kalan_oran * brut_kalan
        net_getiri = self.realize_net + self.kalan_oran * net_kalan

        return {
            "Hisse": self.symbol,
            "Giris Tarihi": self.entry_date.strftime("%Y-%m-%d"),
            "Cikis Tarihi": exit_date.strftime("%Y-%m-%d"),
            "Giris Fiyati": round(self.entry_price, 4),
            "Cikis Fiyati": round(exit_price, 4),
            "Giris Skoru": self.giris_skoru,
            "Giris Sinyali": self.giris_sinyali,
            "Giris RSI": self.giris_rsi,
            "Giris ADX": self.giris_adx,
            "+DI": self.giris_plus_di,
            "-DI": self.giris_minus_di,
            "SMA200 Egim": self.giris_sma200_egim,
            "Giris Hacim Orani": self.giris_hacim,
            "Giris MACD": self.giris_macd,
            "Giris MACD Sinyal": self.giris_macd_sinyal,
            "Giris ATR %": self.giris_atr,
            "20G Breakout": self.giris_breakout,
            "BB Ust Banda Yakin": self.giris_bb_yakin,
            "Kismi Kar Alindi": "EVET" if self.kismi_kar_alindi else "HAYIR",
            "Cikis Nedeni": exit_reason,
            "Bekleme Gunu": self.bar_sayisi,
            "Brut Getiri %": round(brut_getiri * 100, 2),
            "Net Getiri %": round(net_getiri * 100, 2),
        }


# =========================================================
# 7. HİSSE BAZLI BACKTEST
# =========================================================
def backtest_hisse(
    kod: str,
    data_fetcher: DataFetcherProtocol,
    cfg: BacktestConfig,
    baslangic: Optional[str] = None,
    bitis: Optional[str] = None,
) -> pd.DataFrame:
    baslangic = baslangic or cfg.backtest_start
    bitis = bitis or cfg.backtest_end or datetime.now().strftime("%Y-%m-%d")

    try:
        df = data_fetcher.get_ohlcv(kod, start=cfg.data_start, end=bitis)
    except Exception as e:
        raise ValueError(f"Veri çekme hatası ({kod}): {str(e)}")

    try:
        df = indikatorleri_hazirla(df, cfg)
    except Exception as e:
        raise ValueError(f"Indikator hatasi ({kod}): {str(e)}")

    baslangic_ts = pd.Timestamp(baslangic)
    if getattr(df.index, "tz", None) is not None:
        baslangic_ts = baslangic_ts.tz_localize(df.index.tz)
    df = df[df.index >= baslangic_ts]
    if len(df) < 2:
        raise ValueError(f"Backtest donemi icin yeterli veri yok ({kod})")

    islemler = []
    position = None

    for i in range(len(df) - 1):
        bugun = df.iloc[i]
        yarin = df.iloc[i + 1]

        if position is None:
            if bugun["Sinyal"] in cfg.giris_sinyalleri:
                entry_date = yarin.name
                entry_price = float(yarin["Open"]) * (1 + cfg.slippage)
                position = Position(kod, entry_date, entry_price, bugun, cfg)
            continue

        position.update(bugun)
        position.apply_partial_profit(bugun)

        exit_info = position.check_exit(bugun, yarin)
        if exit_info:
            trade = position.close(
                exit_info["date"], exit_info["price"], exit_info["reason"]
            )
            islemler.append(trade)
            position = None

    if position is not None:
        son = df.iloc[-1]
        exit_price = float(son["Close"]) * (1 - cfg.slippage)
        trade = position.close(df.index[-1], exit_price, "DONEM SONU")
        islemler.append(trade)

    return pd.DataFrame(islemler)


# =========================================================
# PORTFOLIO MODÜLLERİNİ İMPORT ET (BacktestConfig tanımlandıktan sonra)
# =========================================================
from engine.portfolio import PositionManager, CashManager, RiskManager, PortfolioReport


# =========================================================
# 8. PORTFÖY SINIFI (YENİ - MODÜLER)
# =========================================================
class Portfolio:
    def __init__(
        self,
        islemler: pd.DataFrame,
        data_fetcher: DataFetcherProtocol,
        cfg: BacktestConfig,
    ):
        self.config = cfg
        self.data_fetcher = data_fetcher
        self.islemler = islemler.copy()
        self.position_manager = PositionManager(cfg)
        self.cash_manager = CashManager(cfg)
        self.risk_manager = RiskManager(cfg)
        self.report = PortfolioReport(cfg)
        self._prepare_data()
        self._prepare_prices()
        self._precompute_entry_exit_dates()

    def _prepare_data(self):
        x = self.islemler
        x["Giriş Tarihi"] = (
            pd.to_datetime(x["Giriş Tarihi"]).dt.tz_localize(None).dt.normalize()
        )
        x["Çıkış Tarihi"] = (
            pd.to_datetime(x["Çıkış Tarihi"]).dt.tz_localize(None).dt.normalize()
        )
        x = x.sort_values(
            ["Giriş Tarihi", "Giriş Skoru", "Giriş ADX"],
            ascending=[True, False, False],
        ).reset_index(drop=True)
        self.islemler = x

    def _prepare_prices(self):
        hisseler = sorted(self.islemler["Hisse"].dropna().unique())
        fiyatlar = {}
        hatalar = []

        min_t = self.islemler["Giriş Tarihi"].min()
        max_t = self.islemler["Çıkış Tarihi"].max()
        start = (min_t - pd.Timedelta(days=10)).strftime("%Y-%m-%d")
        end = (max_t + pd.Timedelta(days=10)).strftime("%Y-%m-%d")

        logger.info(
            f"PYR mark-to-market fiyatları hazırlanıyor: {len(hisseler)} hisse"
        )
        for i, kod in enumerate(hisseler, 1):
            try:
                df = self.data_fetcher.get_ohlcv(kod, start=start, end=end)
                fiyatlar[kod] = df["Close"].astype(float).sort_index()
            except Exception as e:
                hatalar.append({"Hisse": kod, "Hata": str(e)})
                logger.warning(f"{kod} - {str(e)}")
            if i % 25 == 0 or i == len(hisseler):
                logger.info(f"  {i}/{len(hisseler)}")

        self.fiyatlar = fiyatlar
        self.fiyat_hatalari = pd.DataFrame(hatalar)

        if not fiyatlar:
            raise RuntimeError(
                "Hiçbir hisse için fiyat verisi alınamadı. PYR simülasyonu iptal."
            )

    def _build_date_range(self):
        gunler = sorted(
            {
                pd.Timestamp(g).normalize()
                for s in self.fiyatlar.values()
                for g in s.index
            }
        )
        ilk_giris = self.islemler["Giriş Tarihi"].min()
        son_cikis = self.islemler["Çıkış Tarihi"].max()
        return [g for g in gunler if ilk_giris <= g <= son_cikis]

    def _son_fiyat(self, kod: str, tarih: pd.Timestamp) -> float:
        seri = self.fiyatlar.get(kod)
        if seri is None or seri.empty:
            return np.nan
        deger = seri.asof(tarih)
        return float(deger) if not pd.isna(deger) else np.nan

    def _precompute_entry_exit_dates(self):
        self.entries_by_date = {
            date: group.to_dict("records")
            for date, group in self.islemler.groupby("Giriş Tarihi")
        }
        self.exits_by_date = {
            date: group["Hisse"].tolist()
            for date, group in self.islemler.groupby("Çıkış Tarihi")
        }

    def _process_exits(self, tarih: pd.Timestamp) -> float:
        cikis_nakit = 0.0
        cikacak_hisseler = self.exits_by_date.get(tarih, [])
        for kod in cikacak_hisseler:
            if not self.position_manager.has_position(kod):
                continue
            p = self.position_manager.get_position(kod)
            cikis_degeri = p["Toplam Yatırılan TL"] * (
                1.0 + p["Net Getiri %"] / 100.0
            )
            cikis_nakit += cikis_degeri
            self.position_manager.remove_position(kod, tarih, cikis_degeri)
        return cikis_nakit

    def _process_extra_buys(
        self,
        tarih: pd.Timestamp,
        portfoy_degeri: float,
        kullanilabilir_nakit: float,
    ) -> float:
        nakit = kullanilabilir_nakit
        for kod, p in self.position_manager.get_all_open_positions():
            if not self.position_manager.can_add_extra(kod):
                continue
            px = self._son_fiyat(kod, tarih)
            if not np.isfinite(px):
                continue
            ort_maliyet = p["Toplam Yatırılan TL"] / p["Toplam Lot"]
            kar_orani = (px / ort_maliyet - 1.0) * 100.0
            if kar_orani < self.config.profit_trigger_pct:
                continue
            if not self.risk_manager.can_add_to_position(
                p["Toplam Yatırılan TL"], portfoy_degeri
            ):
                continue
            yatirilacak = self.cash_manager.calculate_extra_buy_amount(
                portfoy_degeri, p["Toplam Yatırılan TL"]
            )
            yatirilacak = min(yatirilacak, nakit)
            lot = int(yatirilacak // px) if px > 0 else 0
            if lot < 1:
                continue
            gercek_yatirim = lot * px
            nakit -= gercek_yatirim
            self.position_manager.add_extra_buy(
                kod, tarih, px, lot, gercek_yatirim, kar_orani
            )
        return nakit

    def _process_entries(
        self,
        tarih: pd.Timestamp,
        portfoy_degeri: float,
        kullanilabilir_nakit: float,
    ) -> float:
        nakit = kullanilabilir_nakit
        adaylar = self.entries_by_date.get(tarih, [])
        for r in adaylar:
            kod = r["Hisse"]
            giris_fiyati = float(r["Giriş Fiyatı"])

            if self.position_manager.has_position(kod):
                self.position_manager.add_skipped_signal(
                    tarih,
                    kod,
                    "Tekrar Sinyal",
                    "Açık pozisyon - günlük PYR tetikleyicisi yönetiyor",
                )
                continue

            if not self.risk_manager.can_open_new_position(
                self.position_manager.get_positions_count()
            ):
                self.position_manager.add_skipped_signal(
                    tarih,
                    kod,
                    "İlk Giriş",
                    f"Maksimum {self.config.max_positions} farklı hisse dolu",
                )
                continue

            hedef = self.cash_manager.calculate_entry_amount(
                portfoy_degeri, self.config.first_entry_ratio
            )
            yatirilacak = min(hedef, nakit)
            lot = int(yatirilacak // giris_fiyati) if giris_fiyati > 0 else 0
            if lot < 1:
                self.position_manager.add_skipped_signal(
                    tarih, kod, "İlk Giriş", "Yetersiz nakit"
                )
                continue

            gercek_yatirim = lot * giris_fiyati
            nakit -= gercek_yatirim
            self.position_manager.add_position(
                kod, r, giris_fiyati, lot, gercek_yatirim
            )
        return nakit

    def _mark_to_market(self, tarih: pd.Timestamp) -> float:
        acik_deger = 0.0
        for kod, p in self.position_manager.get_all_open_positions():
            px = self._son_fiyat(kod, tarih)
            acik_deger += (
                p["Toplam Lot"] * px if np.isfinite(px) else p["Toplam Yatırılan TL"]
            )
        return acik_deger

    def run(self) -> Tuple:
        gunler = self._build_date_range()
        for tarih in gunler:
            cikis_nakit = self._process_exits(tarih)
            kullanilabilir_nakit = self.cash_manager.get_cash()

            acik_deger = self._mark_to_market(tarih)
            portfoy_degeri = kullanilabilir_nakit + acik_deger + cikis_nakit

            kullanilabilir_nakit = self._process_extra_buys(
                tarih, portfoy_degeri, kullanilabilir_nakit
            )
            kullanilabilir_nakit = self._process_entries(
                tarih, portfoy_degeri, kullanilabilir_nakit
            )

            self.cash_manager.nakit = kullanilabilir_nakit + cikis_nakit

            acik_piyasa_degeri = self._mark_to_market(tarih)
            portfoy = self.cash_manager.get_cash() + acik_piyasa_degeri
            self.report.append_daily(
                tarih,
                self.cash_manager.get_cash(),
                self.position_manager.get_positions_count(),
                acik_piyasa_degeri,
                portfoy,
            )

        kapali_df = self.position_manager.get_closed_positions_df()
        ek_df = self.position_manager.get_extra_buys_df()
        atlanan_df = self.position_manager.get_skipped_signals_df()
        return self.report.build_results(kapali_df, ek_df, atlanan_df, self.islemler)


# =========================================================
# 9. PORTFÖY SİMÜLASYONU (ana fonksiyon)
# =========================================================
def pyr_portfoy_simulasyonu(
    islemler: pd.DataFrame,
    data_fetcher: DataFetcherProtocol,
    cfg: BacktestConfig,
) -> Tuple:
    if islemler.empty:
        return tuple(pd.DataFrame() for _ in range(7))
    portfolio = Portfolio(islemler, data_fetcher, cfg)
    return portfolio.run()


# =========================================================
# 10. RAPORLAMA FONKSİYONLARI (tam gövdeler)
# =========================================================
def profit_factor(getiriler):
    g = pd.Series(getiriler, dtype=float)
    kazanc = g[g > 0].sum()
    zarar = abs(g[g <= 0].sum())
    return kazanc / zarar if zarar > 0 else np.inf


def maksimum_dusus(getiriler):
    g = pd.Series(getiriler, dtype=float) / 100.0
    sermaye = (1 + g).cumprod()
    zirve = sermaye.cummax()
    dd = (sermaye / zirve - 1) * 100
    return round(abs(float(dd.min())), 2) if len(dd) else 0.0


def _sembol_normalize(deger):
    if pd.isna(deger):
        return ""
    kod = str(deger).strip().upper()
    if kod.endswith(".IS"):
        kod = kod[:-3]
    return kod


def _dogru_deger(deger):
    if pd.isna(deger):
        return False
    if isinstance(deger, (bool, np.bool_)):
        return bool(deger)
    if isinstance(deger, (int, float, np.integer, np.floating)):
        return float(deger) != 0.0
    metin = str(deger).strip().upper()
    return metin in {"1", "TRUE", "EVET", "E", "VAR", "YES", "Y", "X", "+"}


def endeks_listelerini_ozetten_oku(df_ozet):
    """Excel OZET sayfasından endeks üyeliklerini okur."""
    if df_ozet.empty:
        raise ValueError("OZET sayfası boş.")

    kolon_haritasi = {str(c).strip().casefold(): c for c in df_ozet.columns}
    hisse_kolonu = None
    for aday in ("hisse", "sembol", "symbol", "kod", "ticker"):
        if aday in kolon_haritasi:
            hisse_kolonu = kolon_haritasi[aday]
            break
    if hisse_kolonu is None:
        raise ValueError("OZET sayfasında 'Hisse' sütunu bulunamadı.")

    kod_serisi = df_ozet[hisse_kolonu].map(_sembol_normalize)
    gecerli = kod_serisi.ne("")
    kodlar = kod_serisi[gecerli].drop_duplicates().tolist()
    calisma = df_ozet.loc[gecerli].copy()
    calisma["__KOD__"] = kod_serisi[gecerli].values

    b30 = set()
    b50 = set()
    b100 = set()

    sinif_kolonu = None
    for aday in (
        "endeks üyeliği",
        "endeks uyeligi",
        "endeks",
        "index membership",
        "index",
        "üyelik",
        "uyelik",
    ):
        if aday in kolon_haritasi:
            sinif_kolonu = kolon_haritasi[aday]
            break

    if sinif_kolonu is not None:
        for kod, deger in zip(calisma["__KOD__"], calisma[sinif_kolonu]):
            etiket = "" if pd.isna(deger) else str(deger).strip().upper()
            sade = etiket.replace(" ", "").replace("_", "").replace("-", "")
            if any(x in sade for x in ("BIST30", "XU030", "XU30")):
                b30.add(kod)
            elif any(x in sade for x in ("BIST50", "XU050", "XU50")):
                b50.add(kod)
            elif any(x in sade for x in ("BIST100", "XU100")):
                b100.add(kod)

    kolon_eslesmeleri = {
        "b30": ("bist30", "bist 30", "xu030", "xu30"),
        "b50": ("bist50", "bist 50", "xu050", "xu50"),
        "b100": ("bist100", "bist 100", "xu100"),
    }
    bulunan_uyelik_kolonu = False
    for hedef, adaylar in kolon_eslesmeleri.items():
        kolon = next(
            (kolon_haritasi[a.casefold()] for a in adaylar if a.casefold() in kolon_haritasi),
            None,
        )
        if kolon is None:
            continue
        bulunan_uyelik_kolonu = True
        secilenler = set(
            calisma.loc[calisma[kolon].map(_dogru_deger), "__KOD__"]
        )
        if hedef == "b30":
            b30.update(secilenler)
        elif hedef == "b50":
            b50.update(secilenler)
        else:
            b100.update(secilenler)

    bilgi_var = sinif_kolonu is not None or bulunan_uyelik_kolonu
    if not bilgi_var:
        logger.warning(
            "OZET sayfasında endeks üyeliği sütunu bulunamadı. BIST30/50/100 özetleri boş bırakılacak."
        )
    else:
        logger.info(
            f"Endeks üyelikleri okundu | BIST30:{len(b30)} | BIST50:{len(b50)} | BIST100:{len(b100)}"
        )

    return kodlar, b100, b50, b30


def endeks_uyeligi(
    kod: str,
    b100: Set[str],
    b50: Set[str],
    b30: Set[str],
    tum_kodlar: Optional[Set[str]] = None,
) -> str:
    kod = _sembol_normalize(kod)
    if kod in b30:
        return "BIST30"
    if kod in b50:
        return "BIST50"
    if kod in b100:
        return "BIST100"
    if tum_kodlar is not None and kod not in tum_kodlar:
        return "BİLİNMEYEN"
    return "BIST TÜM"


def grup_istatistikleri(grup, ad_sutunu, ad):
    if grup.empty:
        return {
            ad_sutunu: ad,
            "İşlem": 0,
            "Başarı_Oranı": np.nan,
            "Ortalama_Getiri": np.nan,
            "Medyan_Getiri": np.nan,
            "Toplam_Getiri": 0,
        }
    g = grup["Net Getiri %"].astype(float)
    return {
        ad_sutunu: ad,
        "İşlem": len(grup),
        "Başarı_Oranı": round((g > 0).mean() * 100, 2),
        "Ortalama_Getiri": round(g.mean(), 2),
        "Medyan_Getiri": round(g.median(), 2),
        "Toplam_Getiri": round(g.sum(), 2),
    }


def hisse_ozeti_hazirla(islemler):
    satirlar = []
    for kod, grup in islemler.groupby("Hisse"):
        g = grup["Net Getiri %"].astype(float)
        pf = profit_factor(g)
        satirlar.append(
            {
                "Hisse": kod,
                "Toplam İşlem": len(grup),
                "Kazanan": int((g > 0).sum()),
                "Kaybeden": int((g <= 0).sum()),
                "Başarı Oranı %": round((g > 0).mean() * 100, 2),
                "Bileşik Net Getiri %": round(
                    ((1 + g / 100).prod() - 1) * 100, 2
                ),
                "Ortalama İşlem %": round(g.mean(), 2),
                "Medyan İşlem %": round(g.median(), 2),
                "En İyi İşlem %": round(g.max(), 2),
                "En Kötü İşlem %": round(g.min(), 2),
                "Profit Factor": round(pf, 3) if np.isfinite(pf) else "∞",
                "Maksimum Düşüş %": maksimum_dusus(g),
                "Ortalama Bekleme": round(grup["Bekleme Günü"].mean(), 1),
                "Endeks Üyeliği": grup["Endeks Üyeliği"].iloc[0],
            }
        )
    if not satirlar:
        return pd.DataFrame()
    return pd.DataFrame(satirlar).sort_values(
        "Bileşik Net Getiri %", ascending=False
    )


def genel_ozet_hazirla(islemler, bitis_tarihi, cfg: BacktestConfig):
    g = islemler["Net Getiri %"].astype(float)
    pf = profit_factor(g)
    return pd.DataFrame(
        [
            ("Backtest Başlangıcı", cfg.backtest_start),
            ("Backtest Bitişi", bitis_tarihi),
            ("Evren", "BIST TÜM (XUTUM)"),
            ("İşlem Oluşan Hisse Sayısı", islemler["Hisse"].nunique()),
            ("Toplam İşlem", len(islemler)),
            ("Başarı Oranı %", round((g > 0).mean() * 100, 2)),
            ("Ortalama İşlem %", round(g.mean(), 2)),
            ("Medyan İşlem %", round(g.median(), 2)),
            (
                "Teorik Bileşik Getiri %",
                round(((1 + g / 100).prod() - 1) * 100, 2),
            ),
            ("Profit Factor", round(pf, 3) if np.isfinite(pf) else "∞"),
            ("En İyi İşlem %", round(g.max(), 2)),
            ("En Kötü İşlem %", round(g.min(), 2)),
            (
                "Ortalama Bekleme Günü",
                round(islemler["Bekleme Günü"].mean(), 1),
            ),
        ],
        columns=["Açıklama", "Değer"],
    )


def endeks_ozeti_hazirla(islemler):
    gruplar = [
        ("BIST 30", islemler[islemler["Endeks Üyeliği"] == "BIST30"]),
        (
            "BIST 50",
            islemler[islemler["Endeks Üyeliği"].isin(["BIST30", "BIST50"])],
        ),
        (
            "BIST 100",
            islemler[
                islemler["Endeks Üyeliği"].isin(["BIST30", "BIST50", "BIST100"])
            ],
        ),
        ("BIST TÜM", islemler),
    ]
    rows = []
    for ad, grup in gruplar:
        if grup.empty:
            rows.append(
                {
                    "Endeks": ad,
                    "İşlem Oluşan Hisse": 0,
                    "Toplam İşlem": 0,
                    "Başarı Oranı %": np.nan,
                    "Ortalama İşlem %": np.nan,
                    "Medyan İşlem %": np.nan,
                    "Profit Factor": np.nan,
                    "En İyi İşlem %": np.nan,
                    "En Kötü İşlem %": np.nan,
                    "Ortalama Bekleme Günü": np.nan,
                }
            )
            continue
        g = grup["Net Getiri %"].astype(float)
        pf = profit_factor(g)
        rows.append(
            {
                "Endeks": ad,
                "İşlem Oluşan Hisse": grup["Hisse"].nunique(),
                "Toplam İşlem": len(grup),
                "Başarı Oranı %": round((g > 0).mean() * 100, 2),
                "Ortalama İşlem %": round(g.mean(), 2),
                "Medyan İşlem %": round(g.median(), 2),
                "Profit Factor": round(pf, 3) if np.isfinite(pf) else "∞",
                "En İyi İşlem %": round(g.max(), 2),
                "En Kötü İşlem %": round(g.min(), 2),
                "Ortalama Bekleme Günü": round(
                    grup["Bekleme Günü"].mean(), 1
                ),
            }
        )
    return pd.DataFrame(rows)


def yillik_ozet_hazirla(islemler):
    temp = islemler.copy()
    temp["Yıl"] = pd.to_datetime(temp["Çıkış Tarihi"]).dt.year
    rows = []
    for yil, grup in temp.groupby("Yıl"):
        g = grup["Net Getiri %"].astype(float)
        rows.append(
            {
                "Yıl": int(yil),
                "İşlem": len(grup),
                "Başarı_Oranı": round((g > 0).mean() * 100, 2),
                "Ortalama_Getiri": round(g.mean(), 2),
                "Medyan_Getiri": round(g.median(), 2),
                "Toplam_Net_Getiri": round(g.sum(), 2),
            }
        )
    return pd.DataFrame(rows).sort_values("Yıl")


def skor_analizi_hazirla(islemler):
    s = islemler["Giriş Skoru"].astype(float)
    dilimler = [
        ("<85", s < 85),
        ("85-89", (s >= 85) & (s < 90)),
        ("90-94", (s >= 90) & (s < 95)),
        ("95-100", s >= 95),
    ]
    return pd.DataFrame(
        [
            grup_istatistikleri(islemler[m], "Skor Dilimi", ad)
            for ad, m in dilimler
        ]
    )


def cikis_analizi_hazirla(islemler):
    rows = []
    for neden, grup in islemler.groupby("Çıkış Nedeni"):
        g = grup["Net Getiri %"].astype(float)
        rows.append(
            {
                "Çıkış Nedeni": neden,
                "İşlem": len(grup),
                "Başarı_Oranı": round((g > 0).mean() * 100, 2),
                "Ortalama_Getiri": round(g.mean(), 2),
                "Medyan_Getiri": round(g.median(), 2),
                "Toplam_Getiri": round(g.sum(), 2),
                "Ortalama_Bekleme": round(grup["Bekleme Günü"].mean(), 1),
            }
        )
    return pd.DataFrame(rows).sort_values("İşlem", ascending=False)


def faktor_analizi_hazirla(islemler):
    rows = []
    rsi = islemler["Giriş RSI"].astype(float)
    adx = islemler["Giriş ADX"].astype(float)
    hacim = islemler["Giriş Hacim Oranı"].astype(float)

    faktorler = [
        ("RSI", "<55", rsi < 55),
        ("RSI", "55-59.99", (rsi >= 55) & (rsi < 60)),
        ("RSI", "60-64.99", (rsi >= 60) & (rsi < 65)),
        ("RSI", "65-70", (rsi >= 65) & (rsi <= 70)),
        ("RSI", ">70", rsi > 70),
        ("ADX", "<30", adx < 30),
        ("ADX", "30-34.99", (adx >= 30) & (adx < 35)),
        ("ADX", "35-39.99", (adx >= 35) & (adx < 40)),
        ("ADX", "40-44.99", (adx >= 40) & (adx < 45)),
        ("ADX", "45+", adx >= 45),
        ("Hacim", "<1.5x", hacim < 1.5),
        ("Hacim", "1.5-1.99x", (hacim >= 1.5) & (hacim < 2.0)),
        ("Hacim", "2-2.99x", (hacim >= 2.0) & (hacim < 3.0)),
        ("Hacim", "3-4.99x", (hacim >= 3.0) & (hacim < 5.0)),
        ("Hacim", "5x+", hacim >= 5.0),
    ]
    for faktor, dilim, mask in faktorler:
        d = grup_istatistikleri(islemler[mask], "Dilim", dilim)
        d = {"Faktör": faktor, **d}
        rows.append(d)

    for faktor, sutun in [
        ("Breakout", "20G Breakout"),
        ("Bollinger", "BB Üst Banda Yakın"),
    ]:
        for deger in ["EVET", "HAYIR"]:
            d = grup_istatistikleri(
                islemler[islemler[sutun] == deger], "Dilim", deger
            )
            rows.append({"Faktör": faktor, **d})

    return pd.DataFrame(rows)[
        [
            "Faktör",
            "Dilim",
            "İşlem",
            "Başarı_Oranı",
            "Ortalama_Getiri",
            "Medyan_Getiri",
            "Toplam_Getiri",
        ]
    ]


def ayarlar_hazirla(cfg: BacktestConfig):
    return pd.DataFrame(
        [
            ("Versiyon", "A003.3 - Modüler (A002 algoritması)"),
            ("Evren", "BIST TÜM (XUTUM)"),
            ("Minimum VIOS Skoru", cfg.strong_buy_min_score),
            ("ADX Aralığı", f"{cfg.adx_strong}-{cfg.adx_upper}"),
            (
                "Hacim Aralığı",
                f"{cfg.volume_high_ratio}x-{cfg.volume_upper_ratio}x",
            ),
            ("Hibrit Filtre", "Hacim >3.0x ise ADX >=40 zorunlu"),
            (
                "Güçlü Al RSI Aralığı",
                f"{cfg.rsi_low}-{cfg.rsi_high}",
            ),
            ("MACD Filtresi", "MACD > Signal"),
            ("+DI Filtresi", "+DI > -DI zorunlu"),
            ("SMA200 Eğimi", "Pozitif zorunlu"),
            (
                "Breakout Filtresi",
                "Önceki 20 günlük zirve veya BB üst banda %2 yakınlık",
            ),
            ("Giriş", "GÜÇLÜ AL sinyali sonrası ertesi gün açılışı"),
            ("Çıkış 1", f"%{cfg.stop_loss_ratio*100:.0f} stop loss"),
            (
                "Çıkış 2",
                f"+%{cfg.partial_profit_threshold*100:.0f} kârda pozisyonun %{cfg.partial_profit_ratio*100:.0f} kısmını sat",
            ),
            (
                "Çıkış 3",
                f"Kâr %{cfg.trailing_start*100:.0f} sonrası zirveden %{cfg.trailing_stop_ratio*100:.0f} trailing stop",
            ),
            (
                "Çıkış 4",
                "SAT / KÂR AL sinyali sonrası ertesi gün açılışı",
            ),
            (
                "Çıkış 5",
                "SMA50 – 2*ATR altına 1 kapanış sonrası ertesi gün açılışı",
            ),
            ("Çıkış 6", f"{cfg.max_hold_days} işlem günü azami bekleme"),
            ("Komisyon", cfg.commission),
            ("Slippage", cfg.slippage),
            (
                "Not",
                "Temettü, bedelli/bedelsiz ve vergi ayrıca modellenmemiştir.",
            ),
        ],
        columns=["Parametre", "Değer"],
    )


# =========================================================
# 11. EXCEL RAPORLAYICI (eksiksiz)
# =========================================================
class ExcelReporter:
    def __init__(self, filepath):
        self.filepath = filepath
        self.sheets = []

    def add_sheet(self, df, sheet_name, index=False):
        self.sheets.append((df, sheet_name, index))

    @staticmethod
    def _apply_style_to_workbook(workbook):
        header_fill = PatternFill("solid", fgColor="1F4E78")
        alt_fill = PatternFill("solid", fgColor="D9EAF7")
        white_bold = Font(color="FFFFFF", bold=True)

        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for c in ws[1]:
                c.fill = header_fill
                c.font = white_bold
            for row in range(2, ws.max_row + 1):
                if row % 2 == 0:
                    for c in ws[row]:
                        c.fill = alt_fill
            for col in ws.columns:
                mx = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                    max(mx + 2, 11), 34
                )

    def save(self):
        with pd.ExcelWriter(self.filepath, engine="xlsxwriter") as writer:
            for df, sheet_name, index in self.sheets:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        try:
            wb = load_workbook(self.filepath)
            self._apply_style_to_workbook(wb)
            wb.save(self.filepath)
        except Exception as e:
            logger.error(f"Stil uygulama hatası: {str(e)}")


# =========================================================
# 12. SIM001 EKLEME
# =========================================================
def sim001_excel_ekle(
    dosya, islemler, data_fetcher: DataFetcherProtocol, cfg: BacktestConfig
):
    x = islemler.copy()
    x = x.rename(
        columns={
            "Giris Tarihi": "Giriş Tarihi",
            "Cikis Tarihi": "Çıkış Tarihi",
            "Giris Fiyati": "Giriş Fiyatı",
            "Cikis Fiyati": "Çıkış Fiyatı",
            "Giris Skoru": "Giriş Skoru",
            "Giris Sinyali": "Giriş Sinyali",
            "Giris RSI": "Giriş RSI",
            "Giris ADX": "Giriş ADX",
            "Giris Hacim Orani": "Giriş Hacim Oranı",
            "Giris MACD": "Giriş MACD",
            "Giris MACD Sinyal": "Giriş MACD Sinyal",
            "Giris ATR %": "Giriş ATR %",
            "BB Ust Banda Yakin": "BB Üst Banda Yakın",
            "Cikis Nedeni": "Çıkış Nedeni",
            "Bekleme Gunu": "Bekleme Günü",
            "Brut Getiri %": "Brüt Getiri %",
        }
    )
    x = x.sort_values(["Giriş Tarihi", "Hisse"])

    try:
        sonuc = pyr_portfoy_simulasyonu(x, data_fetcher, cfg)
    except Exception as e:
        logger.error(f"SIM001 simülasyonu başarısız: {str(e)}")
        return

    if len(sonuc) < 7:
        logger.error("SIM001 simülasyonu başarısız.")
        return

    kapali, ek, gunluk, yillik, atlanan, fiyat_hatalari, ozet = sonuc

    try:
        with pd.ExcelWriter(
            dosya, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as w:
            ozet.to_excel(w, sheet_name="SIM001 Özet", index=False)
            kapali.to_excel(w, sheet_name="Portföy İşlemleri", index=False)
            ek.to_excel(w, sheet_name="Ek Alımlar", index=False)
            gunluk.to_excel(w, sheet_name="MTM", index=False)
            yillik.to_excel(w, sheet_name="Yıllık", index=False)
    except Exception as e:
        logger.error(f"SIM001 Excel yazma hatası: {str(e)}")
        return

    try:
        wb = load_workbook(dosya)
        ExcelReporter._apply_style_to_workbook(wb)
        wb.save(dosya)
    except Exception as e:
        logger.error(f"SIM001 stil uygulama hatası: {str(e)}")

    print("\nVIOS SIM PORTFÖY BACKTEST 001")
    print("VIOS_SIM_Portfoy_Backtest_003N | A004 | Tetik +%8 | Stop -%10.25")
    print(ozet.to_string(index=False))